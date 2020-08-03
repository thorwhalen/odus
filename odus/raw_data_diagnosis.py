from functools import partial
import os

import pandas as pd
import numpy as np
from openpyxl import load_workbook  # Get it here: https://pypi.org/project/openpyxl/ (or pip install openpyxl)


def pra(obj):
    print(*(x for x in dir(obj) if not x[0].startswith('_') and not x[0].isupper()), sep='\t')


def last_true_idx(series):
    return len(series) - series.iloc[::-1].values.argmax()


def mk_dotpath_getter(dotpath):
    attrs = dotpath.split('.')

    def attr_get(x):
        for a in attrs:
            x = getattr(x, a)
        return x

    return attr_get


def dfof(sheet, dotpath='value'):
    attr_get = mk_dotpath_getter(dotpath)
    data = [[attr_get(c) for c in row] for row in sheet]
    return pd.DataFrame(data)


def coord_val_and_bgcolor_of_cell(c):
    return {'coordinate': c.coordinate, 'value': c.value, 'rgb': c.fill.bgColor.rgb}


def crop(df, last_row_idx, last_col_idx):
    return df.iloc[:last_row_idx, :last_col_idx]


def matmask(df, mask):
    if isinstance(mask, pd.DataFrame):
        mask = mask.values
    d = df.values.copy()
    d[~mask] = ''
    return pd.DataFrame(d, index=df.index, columns=df.columns)


def row_lidx_to_mat_lidx(row_lidx, n_cols):
    return np.tile(row_lidx, (n_cols, 1)).T


white = '00000000'


def valid_vals_int_coords(vals_df):
    first_column = vals_df.iloc[:, 0]
    last_row_idx = last_true_idx((first_column != 'List of States') & (~first_column.isna()))

    age_row = vals_df.iloc[2, :]
    last_col_idx = last_true_idx(~age_row.isna())

    return last_row_idx, last_col_idx


def sheet_to_format_prepped_df(sheet):
    vals = dfof(sheet, dotpath='value')
    color = dfof(sheet, dotpath='fill.bgColor.rgb')
    coordinate = dfof(sheet, dotpath='coordinate')
    #     print(vals.shape, color.shape, coordinate.shape)

    int_coords = valid_vals_int_coords(vals)
    vals, color, coordinate = list(map(lambda x: crop(x, *int_coords), (vals, color, coordinate)))
    #     print(vals.shape, color.shape, coordinate.shape)

    df = pd.concat((vals, color, coordinate), axis=1, keys=('vals', 'color', 'coordinate'))
    #     df = df[~df.iloc[:, 0].isna()]

    return df


def lidx_of_empty_non_white_or_non_empty_white(df):
    is_null = df.vals.isnull()
    is_white = df.color == white
    return (is_null & ~is_white) | (~is_null & is_white)


def mk_known_exceptions_lidx(vals_df):
    n_rows, n_cols = vals_df.shape
    mk_mat_lidx = partial(row_lidx_to_mat_lidx, n_cols=n_cols)

    lidx = mk_mat_lidx([False] * 3 + [True] * (n_rows - 3))
    lidx[:, 0] = False
    lidx &= mk_mat_lidx(~vals_df.iloc[:, 0].isna())
    lidx &= mk_mat_lidx(~(vals_df.iloc[:, 0] == 'Select States below'))
    return lidx


def diagnosis_lidx(df):
    lidx = lidx_of_empty_non_white_or_non_empty_white(df)
    lidx &= mk_known_exceptions_lidx(df.vals)
    return lidx


def diagnosis_lidx_of_xls_file(xls_filepath, sheetname=None):
    wb = load_workbook(xls_filepath, data_only=True)
    sheetname = sheetname or next(iter(wb.sheetnames))
    sheet = wb[sheetname]

    df = sheet_to_format_prepped_df(sheet)
    lidx = lidx_of_empty_non_white_or_non_empty_white(df)
    lidx &= mk_known_exceptions_lidx(df.vals)
    return df, lidx


def diagnosis_items(source_dir, sheetname=None):
    is_excel = lambda x: x.endswith('.xlsx')
    join_dir = lambda *p: os.path.join(source_dir, *p)
    for xls_filepath in map(join_dir, filter(is_excel, os.listdir(source_dir))):
        filename = os.path.basename(xls_filepath)
        df, lidx = diagnosis_lidx_of_xls_file(xls_filepath, sheetname)
        if lidx.sum().sum() != 0:
            yield filename, df, lidx


def print_diagnosis(source_dir='.', sheetname=None):
    """Print the diagnosis of the raw xls files in `source_dir`.
    Namely, for every file that has a problem, the script will print the name of the problematic files,
    and print the coordinates of the cells that should be checked for problems.

    The problems that are checked for:
    - Cells (in the "data" range) that are (a) empty, but have color or (b) NOT empty, but without color.

    See functions `lidx_of_empty_non_white_or_non_empty_white` and `mk_known_exceptions_lidx` functions for details.
    """
    source_dir = os.path.abspath(os.path.expanduser(source_dir))
    print(f'Diagnosing xls files of "{source_dir}" ...\n\n')

    for filename, df, lidx in diagnosis_items(source_dir, sheetname):
        print(f"The possibly problematic cells for {filename}:")
        t = np.ravel(matmask(df.coordinate, lidx))
        t = t[t != '']
        print(*t, sep=', ')
        print("\n")


if __name__ == '__main__':
    import argh

    argh.dispatch_command(print_diagnosis)
